import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


# Cria uma nova planilha
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dias do Mês"

# Define o mês e o ano desejados
ano = 2024
mes = 11

# Adiciona cabeçalhos
ws['A1'] = "Dia"
ws['B1'] = "Dia da Semana"

# Preenche a planilha com os dias do mês e a primeira letra do dia da semana
data_inicial = datetime(ano, mes, 1)
for i in range(1, 32):
    try:
        data_atual = data_inicial + timedelta(days=i-1)
        ws[f'A{i+1}'] = data_atual.day
        dia_semana = data_atual.strftime('%A')[0]
        ws[f'B{i+1}'] = dia_semana

        # Se for sábado ou domingo, preenche as colunas de 1 a 7
        if data_atual.weekday() >= 5:  # 5 é sábado, 6 é domingo
            for col in range(1, 8):
                ws[f'{get_column_letter(col)}{i+1}'] = f"Coluna {col}"

    except ValueError:
        break

# Salva a planilha
wb.save("dias_do_mes.xlsx")