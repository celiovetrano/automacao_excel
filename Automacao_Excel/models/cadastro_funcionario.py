import re
from openpyxl import Workbook, load_workbook
from models.inicializacao import Inicializacao


class CadastroFuncionario:
    # Função para obter RF no formato correto
    def obter_rf(self):
        while True:
            rf = input("Digite o RF (formato 000.000.0.00/0): ")
            if re.match(r"^\d{3}\.\d{3}\.\d\.\d{2}/\d$", rf):
                return rf
            else:
                print("RF inválido. Por favor, insira no formato 000.000.0.00/0.")

    # Função para carregar dados dos funcionários
    def carregar_funcionarios():
        try:
            wb = load_workbook("servidor.xlsx")
            ws = wb.active
            funcionarios = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Garantir que a linha não esteja vazia
                    funcionarios[row[0]] = row[1]
            return funcionarios
        except FileNotFoundError:
            return {}

    # Função para salvar dados dos funcionários
    def salvar_funcionarios(funcionarios):
        wb = Workbook()
        ws = wb.active
        ws.append(["RF", "Nome"])
        for rf, nome in funcionarios.items():
            ws.append([rf, nome])
        wb.save("servidor.xlsx")

    # Função para adicionar um funcionário
    def adicionar_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            nome = input("Digite o nome do funcionário: ").strip()
            funcionarios[rf] = nome
            self.salvar_funcionarios(funcionarios)
            print(f"Funcionário {nome} cadastrado com sucesso!")
        
            if input("Pressione 'Tab' para retornar ao menu: ") == '\t':
                break
        
        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para consultar funcionários
    def consultar_funcionarios(funcionarios):
        while True:
            Inicializacao.limpar_tela()
            if funcionarios:
                for rf, nome in funcionarios.items():
                    print(f"RF: {rf}, Nome: {nome}")
            else:
                print("Nenhum funcionário cadastrado.")

            if input("Pressione 'Tab' para retornar ao menu: ") == '\t':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para alterar dados de um funcionário
    def alterar_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            if rf in funcionarios:
                nome = input(f"Digite o novo nome para o RF {rf}: ").strip()
                funcionarios[rf] = nome
                self.salvar_funcionarios(funcionarios)
                print(f"Dados do funcionário {rf} atualizados com sucesso!")
            else:
                print("Funcionário não encontrado.")

            if input("Pressione 'Tab' para retornar ao menu: ") == '\t':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()

    # Função para excluir um funcionário
    def excluir_funcionario(self, funcionarios):
        while True:
            Inicializacao.limpar_tela()
            rf = self.obter_rf()
            if rf in funcionarios:
                del funcionarios[rf]
                self.salvar_funcionarios(funcionarios)
                print(f"Funcionário {rf} excluído com sucesso!")
            else:
                print("Funcionário não encontrado.")

            if input("Pressione 'Tab' para retornar ao menu: ") == '\t':
                break

        Inicializacao.limpar_tela()
        Inicializacao.exibir_nome_aplicaçao()
        Inicializacao.menu()
