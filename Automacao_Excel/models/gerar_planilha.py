import os
import calendar
from openpyxl import load_workbook
import locale
import datetime

class GerarPlanilha:
    # Função para obter o mês por extenso
    def obter_nome_mes(self, mes):
        # Configurações do usuário
        locale.setlocale(locale.LC_ALL, "")
        nome_mes = calendar.month_name[mes]
        return nome_mes

    # função para recuperar primeiro e último dia do mês / ano
    def primeiro_e_ultimo_dia_do_mes(self, ano, mes):
        # Primeiro dia do mês
        primeiro_dia = datetime.date(ano, mes, 1)
        
        # Último dia do mês
        if mes == 12:
            ultimo_dia = datetime.date(ano + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            ultimo_dia = datetime.date(ano, mes + 1, 1) - datetime.timedelta(days=1)
        
        # Formatar as datas para o padrão brasileiro
        primeiro_dia_formatado = primeiro_dia.strftime('%d/%m/%Y')
        ultimo_dia_formatado = ultimo_dia.strftime('%d/%m/%Y')
        
        return primeiro_dia_formatado, ultimo_dia_formatado

    # validar datas do mês para definir se é final de semana ou está em período de recesso.
    def ajustar_data(self, mes, ano, ws, recesso_inicio=None, recesso_fim=None):
        # Define os dias da semana
        dias_semana_completo = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]
        dias_semana_inicial = ["S", "T", "Q", "Q", "S", "S", "D"]
        
        # Obtem o número de dias no mês e o primeiro dia da semana
        num_dias = calendar.monthrange(ano, mes)[1]
        
        # Define a linha inicial e a coluna inicial
        linha_inicial = 17
        coluna_inicial = 2
        
        for dia in range(1, 32):  # Verificamos até o dia 31
            if dia <= num_dias:
                dia_semana = calendar.weekday(ano, mes, dia)
                ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial, value=dia)  # Preenche o dia
                ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial + 1, value=dias_semana_inicial[dia_semana])  # Preenche a inicial do dia da semana

                # Verifica se o dia está dentro do período de recesso
                if recesso_inicio and recesso_fim and recesso_inicio <= dia <= recesso_fim:
                    # Preenche a coluna V com "Recesso" para os dias do recesso
                    if dia_semana in [5, 6]:  # Mantém a configuração atual para finais de semana
                        ws.cell(row=linha_inicial + dia - 1, column=22, value=dias_semana_completo[dia_semana])  # Coluna V é a coluna 22
                    else:
                        ws.cell(row=linha_inicial + dia - 1, column=22, value="Recesso")
                else:
                    # Preenche a coluna V com o nome do dia da semana por extenso apenas para finais de semana
                    if dia_semana in [5, 6]:
                        ws.cell(row=linha_inicial + dia - 1, column=22, value=dias_semana_completo[dia_semana])  # Coluna V é a coluna 22
                    else:
                        ws.cell(row=linha_inicial + dia - 1, column=22, value=None)  # Limpa a coluna V se não for sábado ou domingo

    def editar_aba(self, arquivo_origem, nome_nova_aba, mes, ano, primeiro_nome, rf_numeros, diretorio, recesso_inicio=None, recesso_fim=None):
        # Abre o arquivo de origem com opção data_only
        wb_origem = load_workbook(arquivo_origem, keep_vba=True, data_only=True)

        # Seleciona a aba ativa
        aba_ativa = wb_origem.active

        # Renomeia a aba ativa
        aba_ativa.title = nome_nova_aba

        # Faz as edições necessárias na aba ativa
        aba_ativa['A1'] = f"Planilha do mês {nome_nova_aba}"

        #obter primeiro e último dia do mês
        primeiro_dia, ultimo_dia = self.primeiro_e_ultimo_dia_do_mes(ano, mes)
        
        # Adiciona o mês e ano na célula A13
        aba_ativa['A13'] = f"{self.obter_nome_mes(mes)}: {primeiro_dia} à {ultimo_dia}".capitalize()

        # Ajusta os dados na aba ativa
        self.ajustar_data(mes, ano, aba_ativa, recesso_inicio, recesso_fim)

        # Garante que o novo arquivo é salvo com a extensão correta
        arquivo_nome = f"planilha_{mes:02}_{ano}_{primeiro_nome}_{rf_numeros}.xlsm"
        wb_origem.save(os.path.join(diretorio, arquivo_nome))
        print(f"Planilha de horas para {primeiro_nome} (RF: {rf_numeros}) gerada com sucesso!")
 
    # Função para validar entrada de dados
    def obter_inteiro_valido(self, prompt, minimo, maximo):
        while True:
            try:
                valor = int(input(prompt))
                if minimo <= valor <= maximo:
                    return valor
                else:
                    print(f"Por favor, insira um valor entre {minimo} e {maximo}.")
            except ValueError:
                print("Entrada inválida. Por favor, insira um número inteiro válido.")

    def obter_resposta_sim_nao(self, prompt):
        while True:
            resposta = input(prompt).strip().lower()
            if resposta in ['s', 'n']:
                return resposta
            else:
                print("Entrada inválida. Por favor, digite 's' para sim ou 'n' para não.")
