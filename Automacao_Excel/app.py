import calendar
from openpyxl import Workbook
#from openpyxl import load_workbook

#arquivo = load_workbook("FFI DOCENTES.xlsm")

#atual_aba = arquivo.active
#nova_aba = arquivo.copy_worksheet(atual_aba)

def criar_aba(nova_aba, arquivo):
    arquivo.create_sheet

def criar_planilha(mes, ano):
    # Cria um novo workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active

    # Define os dias da semana
    dias_semana = ["S", "T", "Q", "Q", "S", "S", "D"]

    # Obtem o número de dias no mês e o primeiro dia da semana
    num_dias = calendar.monthrange(ano, mes)[1]

    # Define a linha inicial e a coluna inicial
    linha_inicial = 14
    coluna_inicial = 2

    for dia in range(1, num_dias + 1):
        dia_semana = calendar.weekday(ano, mes, dia)
        ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial, value=dia)  # Preenche o dia
        ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial + 1, value=dias_semana[dia_semana])  # Preenche o dia da semana

        # Se for sábado (5) ou domingo (6), preenche as colunas 1 a 7
        if dia_semana in [5, 6]:
            for col in range(3, 8):
                ws.cell(row=linha_inicial + dia - 1, column=coluna_inicial + col - 1, value="")

    # Salva o arquivo Excel
    wb.save(f"planilha_{mes}_{ano}.xlsx")

# criando a planilha
mes = int(input("Por favor digite o Mês atual com 2 digitos (MM): "))
ano = int(input("Por favor digite o Ano atual com 4 digitos (AAAA): "))
criar_planilha(mes, ano)
#criar_aba(nova_aba, arquivo)
